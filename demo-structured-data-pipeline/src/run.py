import csv
import json
from pathlib import Path

import jsonschema
import openpyxl
import pdfplumber

INPUT_DIR = Path("input")
OUTPUT_DIR = Path("output")
SCHEMA_PATH = INPUT_DIR / "schema.json"

FIELDS = ["employee_id", "name", "email", "department"]


def normalize(value):
    if isinstance(value, str):
        return value.strip()
    return value


def extract_pdf(path: Path) -> list[dict]:
    """Each PDF is a single 'key: value per line' onboarding form."""
    with pdfplumber.open(path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    record = {}
    for line in text.splitlines():
        if ":" not in line:
            continue
        key, _, value = line.partition(":")
        key = key.strip().lower().replace(" ", "_")
        if key in FIELDS:
            record[key] = normalize(value)
    record["_source"] = path.name
    return [record] if record else []


def extract_excel(path: Path) -> list[dict]:
    """First sheet, first row = headers matching the schema field names."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [normalize(str(c.value)).lower() if c.value is not None else "" for c in ws[1]]

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {header: normalize(cell) for header, cell in zip(headers, row) if header in FIELDS}
        if record:
            record["_source"] = path.name
            records.append(record)
    return records


def coerce_employee_id(record: dict) -> dict:
    """Cast numeric-looking employee_id strings to int; leave others alone so
    schema validation reports the type mismatch instead of silently guessing."""
    raw = record.get("employee_id")
    if isinstance(raw, str) and raw.strip().lstrip("-").isdigit():
        record["employee_id"] = int(raw.strip())
    return record


def dedupe(raw_records: list[dict]) -> tuple[list[dict], list[dict]]:
    """Drop exact repeats of the same employee_id, keeping the first occurrence."""
    seen: dict = {}
    deduped, duplicates = [], []
    for record in raw_records:
        key = record.get("employee_id")
        if key in seen:
            duplicates.append({"raw_data": record, "duplicate_of_source": seen[key]})
            continue
        seen[key] = record.get("_source")
        deduped.append(record)
    return deduped, duplicates


def process_pipeline():
    OUTPUT_DIR.mkdir(exist_ok=True)
    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    validator = jsonschema.Draft7Validator(schema)

    raw_records = []
    for pdf_path in sorted(INPUT_DIR.glob("*.pdf")):
        raw_records.extend(extract_pdf(pdf_path))
    for xlsx_path in sorted(INPUT_DIR.glob("*.xlsx")):
        raw_records.extend(extract_excel(xlsx_path))

    total_ingested = len(raw_records)
    raw_records = [coerce_employee_id(r) for r in raw_records]
    deduped, duplicates = dedupe(raw_records)

    valid_records, error_records = [], []
    for record in deduped:
        source = record.pop("_source", None)
        errors = sorted(validator.iter_errors(record), key=lambda e: list(e.path))
        if errors:
            error_records.append({
                "raw_data": record,
                "source": source,
                "error_type": "SchemaValidationError",
                "error_msg": "; ".join(
                    f"{'.'.join(str(p) for p in e.path) or '(root)'}: {e.message}" for e in errors
                ),
            })
        else:
            valid_records.append({**{f: record.get(f) for f in FIELDS}, "source": source})

    # 1. Valid JSON
    (OUTPUT_DIR / "output.json").write_text(
        json.dumps(valid_records, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    # 2. Valid CSV
    csv_path = OUTPUT_DIR / "output.csv"
    if valid_records:
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=[*FIELDS, "source"])
            writer.writeheader()
            writer.writerows(valid_records)
    else:
        csv_path.write_text("", encoding="utf-8")

    # 3. Errors CSV
    errors_csv_path = OUTPUT_DIR / "errors.csv"
    if error_records:
        with open(errors_csv_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["raw_data", "source", "error_type", "error_msg"])
            writer.writeheader()
            writer.writerows(error_records)
    else:
        errors_csv_path.write_text("", encoding="utf-8")

    # 4. Validation report
    report = {
        "total_ingested": total_ingested,
        "duplicates_removed": len(duplicates),
        "processed": len(deduped),
        "valid_count": len(valid_records),
        "error_count": len(error_records),
        "duplicates": duplicates,
        "errors": error_records,
    }
    (OUTPUT_DIR / "validation_report.json").write_text(
        json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    print(
        f"Pipeline finished. Ingested: {total_ingested}, "
        f"Duplicates removed: {len(duplicates)}, "
        f"Valid: {len(valid_records)}, Errors: {len(error_records)}"
    )


if __name__ == "__main__":
    process_pipeline()
